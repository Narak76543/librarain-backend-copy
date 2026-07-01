import io
import calendar
from datetime import datetime, timezone, timedelta, date
from collections import defaultdict
from fastapi import Query, Depends, status, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func

from main import app
from core.db import get_db
from config import configs
from api.auth_user.views import response
from api.auth_user.models import TBL_AUTH_USER
from api.auth_user.security import require_admin

from api.books.models import TBL_BOOK, TBL_STOCK_IN
from api.categories.models import TBL_CATEGORY
from api.orders.models import TBL_ORDER, TBL_ORDER_ITEM

# ==================== EXCEL AND PDF UTILS ====================
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

def get_excel_styles():
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    return {
        "GREEN_FILL": PatternFill("solid", fgColor="059669"),
        "HEADER_FONT": Font(bold=True, color="FFFFFF", name="Arial", size=11),
        "LIGHT_FILL": PatternFill("solid", fgColor="ECFDF5"),
        "NORMAL_FONT": Font(name="Arial", size=10),
        "BOLD_FONT": Font(bold=True, name="Arial", size=10),
        "CENTER": Alignment(horizontal="center", vertical="center"),
        "LEFT": Alignment(horizontal="left", vertical="center"),
        "RIGHT": Alignment(horizontal="right", vertical="center"),
        "BORDER": thin_border
    }

def get_pdf_styles():
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        textColor=colors.HexColor('#059669'),
        spaceAfter=15
    )
    subtitle_style = ParagraphStyle(
        'SubtitleStyle',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=13,
        textColor=colors.HexColor('#1E293B'),
        spaceBefore=15,
        spaceAfter=10
    )
    cell_style = ParagraphStyle(
        'CellStyle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=colors.HexColor('#1E293B')
    )
    cell_header_style = ParagraphStyle(
        'CellHeaderStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=colors.white
    )
    return title_style, subtitle_style, cell_style, cell_header_style

def make_pdf_kpi_table(kpi_data, col_width=120):
    _, _, cell_style, _ = get_pdf_styles()
    kpi_headers = [Paragraph(f"<b>{k}</b>", ParagraphStyle('H', parent=cell_style, alignment=1)) for k in kpi_data.keys()]
    kpi_values = [Paragraph(f"<font color='#059669' size='11'><b>{v}</b></font>", ParagraphStyle('V', parent=cell_style, alignment=1)) for v in kpi_data.values()]
    t = Table([kpi_headers, kpi_values], colWidths=[col_width]*len(kpi_data))
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#F8FAFC')),
        ('BACKGROUND', (0,1), (-1,1), colors.white),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
    ]))
    return t

def make_pdf_table(headers, rows, col_widths):
    _, _, cell_style, cell_header_style = get_pdf_styles()
    table_data = []
    # Headers
    table_data.append([Paragraph(h, cell_header_style) for h in headers])
    # Data
    for r in rows:
        table_data.append([Paragraph(str(cell), cell_style) for cell in r])
    
    t = Table(table_data, colWidths=col_widths)
    t_style = [
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#059669')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]
    for idx in range(1, len(table_data)):
        if idx % 2 == 0:
            t_style.append(('BACKGROUND', (0, idx), (-1, idx), colors.HexColor('#ECFDF5')))
        else:
            t_style.append(('BACKGROUND', (0, idx), (-1, idx), colors.white))
            
    t.setStyle(TableStyle(t_style))
    return t


# ==================== DATA FETCHERS ====================

def fetch_daily_data(target_date_str: str | None, db: Session):
    if target_date_str:
        try:
            target_date = datetime.strptime(target_date_str, "%Y-%m-%d").date()
        except ValueError:
            target_date = datetime.now(timezone.utc).date()
    else:
        target_date = datetime.now(timezone.utc).date()

    start = datetime.combine(target_date, datetime.min.time()).replace(tzinfo=timezone.utc)
    end = start + timedelta(days=1)

    # Base query - orders in date range
    orders = (
        db.query(TBL_ORDER)
        .options(
            joinedload(TBL_ORDER.order_items).joinedload(TBL_ORDER_ITEM.book),
            joinedload(TBL_ORDER.user)
        )
        .filter(
            TBL_ORDER.created_at >= start,
            TBL_ORDER.created_at < end,
            func.lower(TBL_ORDER.status) != "cancelled",
        )
        .all()
    )

    total_orders = len(orders)
    total_revenue = sum(float(o.total) for o in orders)
    total_books_sold = sum(oi.quantity for o in orders for oi in o.order_items)
    total_cost = sum(
        float(oi.cost_price_at_purchase or (oi.book.cost_price if oi.book else None) or float(oi.price_at_purchase) * 0.6) * oi.quantity
        for o in orders for oi in o.order_items
    )
    total_profit = total_revenue - total_cost
    avg_order_value = (total_revenue / total_orders) if total_orders > 0 else 0.0

    # Hourly chart
    hourly = defaultdict(lambda: {"orders": 0, "revenue": 0.0})
    for o in orders:
        hour = o.created_at.strftime("%H:00")
        hourly[hour]["orders"] += 1
        hourly[hour]["revenue"] += float(o.total)
    
    hourly_chart = [
        {"hour": h, "orders": v["orders"], "revenue": f"{v['revenue']:.2f}"}
        for h, v in sorted(hourly.items())
    ]

    # Delivery Method
    delivery_counts = defaultdict(int)
    for o in orders:
        delivery_counts[o.delivery_way or "Pick Up"] += 1
    total_delivery = sum(delivery_counts.values())
    delivery_chart = [
        {
            "method": method,
            "count": count,
            "percentage": round((count / total_delivery) * 100) if total_delivery > 0 else 0
        }
        for method, count in sorted(delivery_counts.items())
    ]

    # Payment Method
    payment_counts = defaultdict(int)
    for o in orders:
        payment_counts[o.payment_method or "COD"] += 1
    total_payment = sum(payment_counts.values())
    payment_chart = [
        {
            "method": method,
            "count": count,
            "percentage": round((count / total_payment) * 100) if total_payment > 0 else 0
        }
        for method, count in sorted(payment_counts.items())
    ]

    # Delivery Partner Chart
    partner_counts = defaultdict(int)
    for o in orders:
        if o.delivery_partner:
            partner_counts[o.delivery_partner] += 1
    delivery_partner_chart = [
        {"partner": partner, "count": count}
        for partner, count in sorted(partner_counts.items(), key=lambda x: (-x[1], x[0]))
    ]

    # Orders Table
    orders_table = []
    for o in orders:
        books_count = sum(oi.quantity for oi in o.order_items)
        cost_val = sum(
            float(oi.cost_price_at_purchase or (oi.book.cost_price if oi.book else None) or float(oi.price_at_purchase) * 0.6) * oi.quantity
            for oi in o.order_items
        )
        total_val = float(o.total)
        profit_val = total_val - cost_val
        
        items_list = []
        for oi in o.order_items:
            items_list.append({
                "book_title": oi.book.title if oi.book else "Unknown",
                "quantity": oi.quantity,
                "unit_price": f"{float(oi.price_at_purchase):.2f}",
                "subtotal": f"{(float(oi.price_at_purchase) * oi.quantity):.2f}"
            })

        orders_table.append({
            "order_id": str(o.id)[:8].upper(),
            "customer": o.user.full_name if o.user else "Unknown",
            "books_count": books_count,
            "items": items_list,
            "delivery_way": o.delivery_way or "Pick Up",
            "delivery_partner": o.delivery_partner or "",
            "payment_method": o.payment_method or "COD",
            "subtotal": f"{sum(float(oi.price_at_purchase) * oi.quantity for oi in o.order_items):.2f}",
            "total": f"{total_val:.2f}",
            "cost": f"{cost_val:.2f}",
            "profit": f"{profit_val:.2f}",
            "profit_margin": f"{(profit_val / total_val * 100):.1f}%" if total_val > 0 else "0%",
            "status": o.status or "pending",
            "created_at": o.created_at.strftime("%I:%M %p")
        })

    # Stock table
    book_sales = defaultdict(lambda: {"qty": 0, "revenue": 0.0, "title": "", "category": "", "remaining": 0, "cost_price": 0.0, "sale_price": 0.0, "has_real_cost": True})
    for o in orders:
        for oi in o.order_items:
            if oi.book:
                bid = str(oi.book_id)
                book_sales[bid]["qty"] += oi.quantity
                book_sales[bid]["revenue"] += float(oi.price_at_purchase) * oi.quantity
                book_sales[bid]["title"] = oi.book.title
                book_sales[bid]["category"] = oi.book.category.name if oi.book.category else "—"
                book_sales[bid]["remaining"] = oi.book.stock
                
                has_real = bool(oi.book.cost_price)
                cp = float(oi.book.cost_price) if oi.book.cost_price else float(oi.book.price) * 0.6
                book_sales[bid]["cost_price"] = cp
                book_sales[bid]["sale_price"] = float(oi.book.price)
                if not has_real:
                    book_sales[bid]["has_real_cost"] = False

    stock_table = []
    for v in book_sales.values():
        cost_today = v["cost_price"] * v["qty"]
        revenue_today = v["revenue"]
        profit_today = revenue_today - cost_today
        margin = (profit_today / revenue_today * 100) if revenue_today > 0 else 0
        stock_table.append({
            "book_title": v["title"],
            "category": v["category"],
            "cost_price": f"{v['cost_price']:.2f}",
            "sale_price": f"{v['sale_price']:.2f}",
            "stock_out_today": v["qty"],
            "remaining_stock": v["remaining"],
            "revenue_today": f"{revenue_today:.2f}",
            "cost_today": f"{cost_today:.2f}",
            "profit_today": f"{profit_today:.2f}",
            "profit_margin": f"{margin:.1f}%",
            "has_real_cost": v["has_real_cost"]
        })

    return {
        "date": target_date.strftime("%Y-%m-%d"),
        "kpis": {
            "total_orders": total_orders,
            "total_revenue": f"{total_revenue:.2f}",
            "total_books_sold": total_books_sold,
            "total_profit": f"{total_profit:.2f}",
            "avg_order_value": f"{avg_order_value:.2f}"
        },
        "hourly_chart": hourly_chart,
        "delivery_chart": delivery_chart,
        "payment_chart": payment_chart,
        "delivery_partner_chart": delivery_partner_chart,
        "orders_table": orders_table,
        "stock_table": stock_table
    }

def fetch_weekly_data(target_date_str: str | None, db: Session):
    if target_date_str:
        try:
            target_date = datetime.strptime(target_date_str, "%Y-%m-%d").date()
        except ValueError:
            target_date = datetime.now(timezone.utc).date()
    else:
        target_date = datetime.now(timezone.utc).date()

    start_of_week = target_date - timedelta(days=target_date.weekday())
    end_of_week = start_of_week + timedelta(days=6)

    start = datetime.combine(start_of_week, datetime.min.time()).replace(tzinfo=timezone.utc)
    end = datetime.combine(end_of_week, datetime.max.time()).replace(tzinfo=timezone.utc)

    orders = (
        db.query(TBL_ORDER)
        .options(
            joinedload(TBL_ORDER.order_items).joinedload(TBL_ORDER_ITEM.book),
            joinedload(TBL_ORDER.user)
        )
        .filter(
            TBL_ORDER.created_at >= start,
            TBL_ORDER.created_at <= end,
            func.lower(TBL_ORDER.status) != "cancelled",
        )
        .all()
    )

    total_orders = len(orders)
    total_revenue = sum(float(o.total) for o in orders)
    total_books_sold = sum(oi.quantity for o in orders for oi in o.order_items if oi.book)
    total_cost = sum(
        float(oi.cost_price_at_purchase or (oi.book.cost_price if oi.book else None) or float(oi.price_at_purchase) * 0.6) * oi.quantity
        for o in orders for oi in o.order_items
    )
    total_profit = total_revenue - total_cost
    avg_daily_revenue = total_revenue / 7.0

    day_names = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    full_day_names = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    
    daily_data = {d: {"orders": 0, "revenue": 0.0, "books": 0, "cost": 0.0} for d in range(7)}
    for o in orders:
        d = o.created_at.weekday()
        if d in daily_data:
            daily_data[d]["orders"] += 1
            daily_data[d]["revenue"] += float(o.total)
            for oi in o.order_items:
                daily_data[d]["books"] += oi.quantity
                daily_data[d]["cost"] += float(oi.cost_price_at_purchase or (oi.book.cost_price if oi.book else None) or float(oi.price_at_purchase) * 0.6) * oi.quantity

    daily_chart = []
    daily_table = []
    for d in range(7):
        day_date = start_of_week + timedelta(days=d)
        rev = daily_data[d]["revenue"]
        prof = rev - daily_data[d]["cost"]
        daily_chart.append({
            "day": day_names[d],
            "orders": daily_data[d]["orders"],
            "revenue": f"{rev:.2f}"
        })
        daily_table.append({
            "date": f"{full_day_names[d]}, {day_date.strftime('%b %d, %Y')}",
            "orders": daily_data[d]["orders"],
            "revenue": f"{rev:.2f}",
            "books_sold": daily_data[d]["books"],
            "profit": f"{prof:.2f}"
        })

    # Delivery Method
    delivery_counts = defaultdict(int)
    for o in orders:
        delivery_counts[o.delivery_way or "Pick Up"] += 1
    total_delivery = sum(delivery_counts.values())
    delivery_chart = [
        {
            "method": method,
            "count": count,
            "percentage": round((count / total_delivery) * 100) if total_delivery > 0 else 0
        }
        for method, count in sorted(delivery_counts.items())
    ]

    # Payment Method
    payment_counts = defaultdict(int)
    for o in orders:
        payment_counts[o.payment_method or "COD"] += 1
    total_payment = sum(payment_counts.values())
    payment_chart = [
        {
            "method": method,
            "count": count,
            "percentage": round((count / total_payment) * 100) if total_payment > 0 else 0
        }
        for method, count in sorted(payment_counts.items())
    ]

    # Stock table
    stock_ins = db.query(TBL_STOCK_IN).filter(
        TBL_STOCK_IN.created_at >= start,
        TBL_STOCK_IN.created_at <= end
    ).all()

    book_activity = defaultdict(lambda: {"stock_in": 0, "stock_out": 0, "revenue": 0.0, "title": "", "category": "", "current_stock": 0, "cost_price": 0.0, "sale_price": 0.0, "has_real_cost": True})
    for o in orders:
        for oi in o.order_items:
            if oi.book:
                bid = oi.book_id
                book_activity[bid]["stock_out"] += oi.quantity
                book_activity[bid]["revenue"] += float(oi.price_at_purchase) * oi.quantity
                book_activity[bid]["title"] = oi.book.title
                book_activity[bid]["category"] = oi.book.category.name if oi.book.category else "—"
                book_activity[bid]["current_stock"] = oi.book.stock
                
                has_real = bool(oi.book.cost_price)
                cp = float(oi.book.cost_price) if oi.book.cost_price else float(oi.book.price) * 0.6
                book_activity[bid]["cost_price"] = cp
                book_activity[bid]["sale_price"] = float(oi.book.price)
                if not has_real:
                    book_activity[bid]["has_real_cost"] = False

    for si in stock_ins:
        if si.book:
            bid = si.book_id
            book_activity[bid]["stock_in"] += si.quantity
            book_activity[bid]["title"] = si.book.title
            book_activity[bid]["category"] = si.book.category.name if si.book.category else "—"
            book_activity[bid]["current_stock"] = si.book.stock
            
            has_real = bool(si.book.cost_price)
            cp = float(si.book.cost_price) if si.book.cost_price else float(si.book.price) * 0.6
            book_activity[bid]["cost_price"] = cp
            book_activity[bid]["sale_price"] = float(si.book.price)
            if not has_real:
                book_activity[bid]["has_real_cost"] = False

    stock_table = []
    for act in book_activity.values():
        closing = act["current_stock"]
        opening = closing + act["stock_out"] - act["stock_in"]
        cost_today = act["cost_price"] * act["stock_out"]
        revenue_today = act["revenue"]
        profit_today = revenue_today - cost_today
        margin = (profit_today / revenue_today * 100) if revenue_today > 0 else 0

        stock_table.append({
            "book_title": act["title"],
            "category": act["category"],
            "cost_price": f"{act['cost_price']:.2f}",
            "sale_price": f"{act['sale_price']:.2f}",
            "opening_stock": opening,
            "stock_in": act["stock_in"],
            "stock_out": act["stock_out"],
            "closing_stock": closing,
            "total_revenue": f"{revenue_today:.2f}",
            "profit": f"{profit_today:.2f}",
            "profit_margin": f"{margin:.1f}%",
            "has_real_cost": act["has_real_cost"]
        })

    week_label = f"{start_of_week.strftime('%b %d, %Y')} - {end_of_week.strftime('%b %d, %Y')}"

    return {
        "week_start": start_of_week.strftime("%Y-%m-%d"),
        "week_end": end_of_week.strftime("%Y-%m-%d"),
        "week_label": week_label,
        "kpis": {
            "total_orders": total_orders,
            "total_revenue": f"{total_revenue:.2f}",
            "total_books_sold": total_books_sold,
            "total_profit": f"{total_profit:.2f}",
            "avg_daily_revenue": f"{avg_daily_revenue:.2f}"
        },
        "daily_chart": daily_chart,
        "delivery_chart": delivery_chart,
        "payment_chart": payment_chart,
        "daily_table": daily_table,
        "stock_table": stock_table
    }

def fetch_monthly_data(month: int | None, year: int | None, db: Session):
    current_time = datetime.now(timezone.utc)
    target_month = month if month is not None else current_time.month
    target_year = year if year is not None else current_time.year

    start_date = datetime(target_year, target_month, 1, 0, 0, 0, tzinfo=timezone.utc)
    last_day = calendar.monthrange(target_year, target_month)[1]
    end_date = datetime(target_year, target_month, last_day, 23, 59, 59, 999999, tzinfo=timezone.utc)

    orders = (
        db.query(TBL_ORDER)
        .options(
            joinedload(TBL_ORDER.order_items).joinedload(TBL_ORDER_ITEM.book),
            joinedload(TBL_ORDER.user)
        )
        .filter(
            TBL_ORDER.created_at >= start_date,
            TBL_ORDER.created_at <= end_date,
            func.lower(TBL_ORDER.status) != "cancelled",
        )
        .all()
    )

    total_orders = len(orders)
    total_revenue = sum(float(o.total) for o in orders)
    total_books_sold = sum(oi.quantity for o in orders for oi in o.order_items if oi.book)
    total_cost = sum(
        float(oi.cost_price_at_purchase or (oi.book.cost_price if oi.book else None) or float(oi.price_at_purchase) * 0.6) * oi.quantity
        for o in orders for oi in o.order_items
    )
    total_profit = total_revenue - total_cost
    avg_daily_revenue = total_revenue / last_day if last_day > 0 else 0.0

    # Best selling book
    book_quantities = defaultdict(int)
    for o in orders:
        for oi in o.order_items:
            if oi.book:
                book_quantities[oi.book.title] += oi.quantity
    best_selling_book = max(book_quantities.keys(), key=lambda k: book_quantities[k]) if book_quantities else "—"

    # Daily trend chart & table
    daily_revenue_chart = []
    daily_table = []
    day_data = {d: {"orders": 0, "revenue": 0.0, "books": 0, "cost": 0.0} for d in range(1, last_day + 1)}
    for o in orders:
        day = o.created_at.day
        if day in day_data:
            day_data[day]["orders"] += 1
            day_data[day]["revenue"] += float(o.total)
            for oi in o.order_items:
                day_data[day]["books"] += oi.quantity
                day_data[day]["cost"] += float(oi.cost_price_at_purchase or (oi.book.cost_price if oi.book else None) or float(oi.price_at_purchase) * 0.6) * oi.quantity

    month_abbr = calendar.month_name[target_month][:3]
    for day in range(1, last_day + 1):
        rev = day_data[day]["revenue"]
        prof = rev - day_data[day]["cost"]
        daily_revenue_chart.append({
            "day": f"{month_abbr} {day}",
            "orders": day_data[day]["orders"],
            "revenue": f"{rev:.2f}"
        })
        daily_table.append({
            "date": f"{month_abbr} {day}, {target_year}",
            "orders": day_data[day]["orders"],
            "revenue": f"{rev:.2f}",
            "books_sold": day_data[day]["books"],
            "profit": f"{prof:.2f}"
        })

    # Category chart
    category_counts = defaultdict(int)
    for o in orders:
        for oi in o.order_items:
            if oi.book and oi.book.category:
                category_counts[oi.book.category.name] += oi.quantity
    total_cat_qty = sum(category_counts.values())
    category_chart = [
        {
            "category": cat,
            "count": count,
            "percentage": round((count / total_cat_qty) * 100) if total_cat_qty > 0 else 0
        }
        for cat, count in sorted(category_counts.items(), key=lambda x: -x[1])
    ]

    # Top books chart
    book_sales_month = defaultdict(lambda: {"qty": 0, "revenue": 0.0, "title": ""})
    for o in orders:
        for oi in o.order_items:
            if oi.book:
                bid = oi.book_id
                book_sales_month[bid]["qty"] += oi.quantity
                book_sales_month[bid]["revenue"] += float(oi.price_at_purchase) * oi.quantity
                book_sales_month[bid]["title"] = oi.book.title
    sorted_books = sorted(book_sales_month.values(), key=lambda x: -x["qty"])[:5]
    top_books_chart = [
        {
            "title": b["title"],
            "qty_sold": b["qty"],
            "revenue": f"{b['revenue']:.2f}"
        }
        for b in sorted_books
    ]

    # Delivery Method
    delivery_counts = defaultdict(int)
    for o in orders:
        delivery_counts[o.delivery_way or "Pick Up"] += 1
    total_delivery = sum(delivery_counts.values())
    delivery_chart = [
        {
            "method": method,
            "count": count,
            "percentage": round((count / total_delivery) * 100) if total_delivery > 0 else 0
        }
        for method, count in sorted(delivery_counts.items())
    ]

    # Payment Method
    payment_counts = defaultdict(int)
    for o in orders:
        payment_counts[o.payment_method or "COD"] += 1
    total_payment = sum(payment_counts.values())
    payment_chart = [
        {
            "method": method,
            "count": count,
            "percentage": round((count / total_payment) * 100) if total_payment > 0 else 0
        }
        for method, count in sorted(payment_counts.items())
    ]

    # Stock table
    stock_ins = db.query(TBL_STOCK_IN).filter(
        TBL_STOCK_IN.created_at >= start_date,
        TBL_STOCK_IN.created_at <= end_date
    ).all()

    book_activity = defaultdict(lambda: {"stock_in": 0, "stock_out": 0, "revenue": 0.0, "title": "", "category": "", "current_stock": 0, "cost_price": 0.0, "sale_price": 0.0, "has_real_cost": True})
    for o in orders:
        for oi in o.order_items:
            if oi.book:
                bid = oi.book_id
                book_activity[bid]["stock_out"] += oi.quantity
                book_activity[bid]["revenue"] += float(oi.price_at_purchase) * oi.quantity
                book_activity[bid]["title"] = oi.book.title
                book_activity[bid]["category"] = oi.book.category.name if oi.book.category else "—"
                book_activity[bid]["current_stock"] = oi.book.stock
                
                has_real = bool(oi.book.cost_price)
                cp = float(oi.book.cost_price) if oi.book.cost_price else float(oi.book.price) * 0.6
                book_activity[bid]["cost_price"] = cp
                book_activity[bid]["sale_price"] = float(oi.book.price)
                if not has_real:
                    book_activity[bid]["has_real_cost"] = False

    for si in stock_ins:
        if si.book:
            bid = si.book_id
            book_activity[bid]["stock_in"] += si.quantity
            book_activity[bid]["title"] = si.book.title
            book_activity[bid]["category"] = si.book.category.name if si.book.category else "—"
            book_activity[bid]["current_stock"] = si.book.stock
            
            has_real = bool(si.book.cost_price)
            cp = float(si.book.cost_price) if si.book.cost_price else float(si.book.price) * 0.6
            book_activity[bid]["cost_price"] = cp
            book_activity[bid]["sale_price"] = float(si.book.price)
            if not has_real:
                book_activity[bid]["has_real_cost"] = False

    stock_table = []
    for act in book_activity.values():
        closing = act["current_stock"]
        opening = closing + act["stock_out"] - act["stock_in"]
        cost_today = act["cost_price"] * act["stock_out"]
        revenue_today = act["revenue"]
        profit_today = revenue_today - cost_today
        margin = (profit_today / revenue_today * 100) if revenue_today > 0 else 0

        stock_table.append({
            "book_title": act["title"],
            "category": act["category"],
            "cost_price": f"{act['cost_price']:.2f}",
            "sale_price": f"{act['sale_price']:.2f}",
            "opening_stock": opening,
            "stock_in": act["stock_in"],
            "stock_out": act["stock_out"],
            "closing_stock": closing,
            "total_revenue": f"{revenue_today:.2f}",
            "profit": f"{profit_today:.2f}",
            "profit_margin": f"{margin:.1f}%",
            "has_real_cost": act["has_real_cost"]
        })

    return {
        "month": target_month,
        "year": target_year,
        "month_name": f"{calendar.month_name[target_month]} {target_year}",
        "kpis": {
            "total_orders": total_orders,
            "total_revenue": f"{total_revenue:.2f}",
            "total_books_sold": total_books_sold,
            "total_profit": f"{total_profit:.2f}",
            "best_selling_book": best_selling_book,
            "avg_daily_revenue": f"{avg_daily_revenue:.2f}"
        },
        "daily_revenue_chart": daily_revenue_chart,
        "category_chart": category_chart,
        "top_books_chart": top_books_chart,
        "delivery_chart": delivery_chart,
        "payment_chart": payment_chart,
        "daily_table": daily_table,
        "stock_table": stock_table
    }

def fetch_yearly_data(year: int | None, db: Session):
    target_year = year if year is not None else datetime.now(timezone.utc).year

    start_date = datetime(target_year, 1, 1, 0, 0, 0, tzinfo=timezone.utc)
    end_date = datetime(target_year, 12, 31, 23, 59, 59, 999999, tzinfo=timezone.utc)

    orders = (
        db.query(TBL_ORDER)
        .options(
            joinedload(TBL_ORDER.order_items).joinedload(TBL_ORDER_ITEM.book),
            joinedload(TBL_ORDER.user)
        )
        .filter(
            TBL_ORDER.created_at >= start_date,
            TBL_ORDER.created_at <= end_date,
            func.lower(TBL_ORDER.status) != "cancelled",
        )
        .all()
    )

    total_orders = len(orders)
    total_revenue = sum(float(o.total) for o in orders)
    total_books_sold = sum(oi.quantity for o in orders for oi in o.order_items if oi.book)
    total_cost = sum(
        float(oi.cost_price_at_purchase or (oi.book.cost_price if oi.book else None) or float(oi.price_at_purchase) * 0.6) * oi.quantity
        for o in orders for oi in o.order_items
    )
    total_profit = total_revenue - total_cost
    avg_monthly_revenue = total_revenue / 12.0

    month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    full_month_names = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
    
    month_data = {m: {"orders": 0, "revenue": 0.0, "books": 0, "cost": 0.0} for m in range(1, 13)}
    for o in orders:
        m = o.created_at.month
        if m in month_data:
            month_data[m]["orders"] += 1
            month_data[m]["revenue"] += float(o.total)
            for oi in o.order_items:
                month_data[m]["books"] += oi.quantity
                month_data[m]["cost"] += float(oi.cost_price_at_purchase or (oi.book.cost_price if oi.book else None) or float(oi.price_at_purchase) * 0.6) * oi.quantity

    best_m_idx = 1
    max_rev = -1.0
    for m in range(1, 13):
        if month_data[m]["revenue"] > max_rev:
            max_rev = month_data[m]["revenue"]
            best_m_idx = m
    best_month = full_month_names[best_m_idx-1] if max_rev > 0 else "—"

    monthly_revenue_chart = []
    monthly_orders_chart = []
    monthly_table = []
    for m in range(1, 13):
        rev = month_data[m]["revenue"]
        prof = rev - month_data[m]["cost"]
        monthly_revenue_chart.append({
            "month": month_names[m-1],
            "orders": month_data[m]["orders"],
            "revenue": f"{rev:.2f}",
            "profit": f"{prof:.2f}"
        })
        monthly_orders_chart.append({
            "month": month_names[m-1],
            "orders": month_data[m]["orders"]
        })
        monthly_table.append({
            "month": f"{full_month_names[m-1]} {target_year}",
            "orders": month_data[m]["orders"],
            "revenue": f"{rev:.2f}",
            "books_sold": month_data[m]["books"],
            "profit": f"{prof:.2f}"
        })

    # Category chart
    category_counts = defaultdict(int)
    for o in orders:
        for oi in o.order_items:
            if oi.book and oi.book.category:
                category_counts[oi.book.category.name] += oi.quantity
    total_cat_qty = sum(category_counts.values())
    category_chart = [
        {
            "category": cat,
            "count": count,
            "percentage": round((count / total_cat_qty) * 100) if total_cat_qty > 0 else 0
        }
        for cat, count in sorted(category_counts.items(), key=lambda x: -x[1])
    ]

    # Top books
    book_sales_year = defaultdict(lambda: {"qty": 0, "revenue": 0.0, "title": "", "remaining": 0})
    for o in orders:
        for oi in o.order_items:
            if oi.book:
                bid = oi.book_id
                book_sales_year[bid]["qty"] += oi.quantity
                book_sales_year[bid]["revenue"] += float(oi.price_at_purchase) * oi.quantity
                book_sales_year[bid]["title"] = oi.book.title
                book_sales_year[bid]["remaining"] = oi.book.stock
    sorted_books_year = sorted(book_sales_year.values(), key=lambda x: -x["qty"])

    top_books_chart = [
        {
            "title": b["title"],
            "qty_sold": b["qty"],
            "revenue": f"{b['revenue']:.2f}"
        }
        for b in sorted_books_year[:10]
    ]

    stock_table = [
        {
            "book_title": b["title"],
            "total_sold": b["qty"],
            "total_revenue": f"{b['revenue']:.2f}",
            "remaining": b["remaining"]
        }
        for b in sorted_books_year
    ]

    return {
        "year": target_year,
        "kpis": {
            "total_orders": total_orders,
            "total_revenue": f"{total_revenue:.2f}",
            "total_books_sold": total_books_sold,
            "total_profit": f"{total_profit:.2f}",
            "avg_monthly_revenue": f"{avg_monthly_revenue:.2f}",
            "best_month": best_month
        },
        "monthly_revenue_chart": monthly_revenue_chart,
        "monthly_orders_chart": monthly_orders_chart,
        "category_chart": category_chart,
        "top_books_chart": top_books_chart,
        "monthly_table": monthly_table,
        "stock_table": stock_table
    }


# ==================== JSON ENDPOINTS ====================

@app.get("/api/v1/admin/reports/daily", tags=["Reports"])
def get_daily_report(
    date: str | None = Query(None),
    current_user: TBL_AUTH_USER = Depends(require_admin),
    db: Session = Depends(get_db)
):
    try:
        report_data = fetch_daily_data(date, db)
        return response(
            ok=True,
            status_code=status.HTTP_200_OK,
            message="Daily report retrieved",
            data=report_data
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to generate daily report: {str(e)}"
        )

@app.get("/api/v1/admin/reports/weekly", tags=["Reports"])
def get_weekly_report_json(
    date: str | None = Query(None),
    current_user: TBL_AUTH_USER = Depends(require_admin),
    db: Session = Depends(get_db)
):
    try:
        report_data = fetch_weekly_data(date, db)
        return response(
            ok=True,
            status_code=status.HTTP_200_OK,
            message="Weekly report retrieved",
            data=report_data
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to generate weekly report: {str(e)}"
        )

@app.get("/api/v1/admin/reports/monthly", tags=["Reports"])
def get_monthly_report(
    month: int | None = Query(None),
    year: int | None = Query(None),
    current_user: TBL_AUTH_USER = Depends(require_admin),
    db: Session = Depends(get_db)
):
    try:
        report_data = fetch_monthly_data(month, year, db)
        return response(
            ok=True,
            status_code=status.HTTP_200_OK,
            message="Monthly report retrieved",
            data=report_data
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to generate monthly report: {str(e)}"
        )

@app.get("/api/v1/admin/reports/yearly", tags=["Reports"])
def get_yearly_report(
    year: int | None = Query(None),
    current_user: TBL_AUTH_USER = Depends(require_admin),
    db: Session = Depends(get_db)
):
    try:
        report_data = fetch_yearly_data(year, db)
        return response(
            ok=True,
            status_code=status.HTTP_200_OK,
            message="Yearly report retrieved",
            data=report_data
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to generate yearly report: {str(e)}"
        )


# ==================== EXPORT ENDPOINTS ====================

@app.get("/api/v1/admin/reports/daily/export", tags=["Reports"])
def export_daily_report(
    date: str | None = Query(None),
    format: str = Query("excel"),
    current_user: TBL_AUTH_USER = Depends(require_admin),
    db: Session = Depends(get_db)
):
    report_data = fetch_daily_data(date, db)
    target_date = report_data["date"]
    
    if format == "excel":
        wb = openpyxl.Workbook()
        style = get_excel_styles()
        
        # --- Sheet 1: Summary ---
        ws1 = wb.active
        ws1.title = "Summary"
        ws1.views.sheetView[0].showGridLines = True
        
        ws1.append(["Daily Report Summary", target_date])
        ws1.cell(1, 1).font = Font(bold=True, size=14, name="Arial")
        ws1.append([])
        
        ws1.append(["KPI Metric", "Value"])
        ws1.cell(3, 1).fill = style["GREEN_FILL"]
        ws1.cell(3, 1).font = style["HEADER_FONT"]
        ws1.cell(3, 2).fill = style["GREEN_FILL"]
        ws1.cell(3, 2).font = style["HEADER_FONT"]
        
        kpi_rows = [
            ("Total Orders", int(report_data["kpis"]["total_orders"])),
            ("Total Revenue", f"${report_data['kpis']['total_revenue']}"),
            ("Total Books Sold", int(report_data["kpis"]["total_books_sold"])),
            ("Net Profit", f"${report_data['kpis']['total_profit']}"),
            ("Average Order Value", f"${report_data['kpis']['avg_order_value']}")
        ]
        for row in kpi_rows:
            ws1.append(row)
            ws1.cell(ws1.max_row, 1).font = style["BOLD_FONT"]
            ws1.cell(ws1.max_row, 1).border = style["BORDER"]
            ws1.cell(ws1.max_row, 2).border = style["BORDER"]
            
        ws1.column_dimensions['A'].width = 25
        ws1.column_dimensions['B'].width = 15
        
        # --- Sheet 2: Orders ---
        ws2 = wb.create_sheet("Orders")
        ws2.views.sheetView[0].showGridLines = True
        
        headers = ["Order ID", "Customer", "Books", "Delivery", "Partner", "Payment", "Total", "Profit", "Status", "Created At"]
        ws2.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws2.cell(1, col_idx)
            cell.fill = style["GREEN_FILL"]
            cell.font = style["HEADER_FONT"]
            cell.alignment = style["CENTER"]
            
        for i, o in enumerate(report_data["orders_table"]):
            row_data = [
                o["order_id"], o["customer"], o["books_count"], o["delivery_way"],
                o["delivery_partner"], o["payment_method"], float(o["total"]),
                float(o["profit"]), o["status"], o["created_at"]
            ]
            ws2.append(row_data)
            curr_row = ws2.max_row
            fill = style["LIGHT_FILL"] if i % 2 == 0 else PatternFill(fill_type=None)
            
            for col_idx in range(1, len(headers) + 1):
                c = ws2.cell(curr_row, col_idx)
                c.border = style["BORDER"]
                c.font = style["NORMAL_FONT"]
                if fill.fill_type:
                    c.fill = fill
                if col_idx in [1, 3, 9, 10]:
                    c.alignment = style["CENTER"]
                elif col_idx in [7, 8]:
                    c.alignment = style["RIGHT"]
                    c.number_format = '"$"#,##0.00'
                    
        column_widths = [12, 20, 10, 15, 12, 12, 12, 12, 12, 15]
        for idx, width in enumerate(column_widths):
            ws2.column_dimensions[openpyxl.utils.get_column_letter(idx + 1)].width = width
            
        # --- Sheet 3: Stock ---
        ws3 = wb.create_sheet("Stock")
        ws3.views.sheetView[0].showGridLines = True
        
        stock_headers = ["Book Title", "Category", "Sold Today", "Remaining Stock", "Revenue"]
        ws3.append(stock_headers)
        for col_idx in range(1, len(stock_headers) + 1):
            cell = ws3.cell(1, col_idx)
            cell.fill = style["GREEN_FILL"]
            cell.font = style["HEADER_FONT"]
            cell.alignment = style["CENTER"]
            
        for i, s in enumerate(report_data["stock_table"]):
            row_data = [
                s["book_title"], s["category"], int(s["stock_out_today"]),
                int(s["remaining_stock"]), float(s["revenue_today"])
            ]
            ws3.append(row_data)
            curr_row = ws3.max_row
            fill = style["LIGHT_FILL"] if i % 2 == 0 else PatternFill(fill_type=None)
            
            for col_idx in range(1, len(stock_headers) + 1):
                c = ws3.cell(curr_row, col_idx)
                c.border = style["BORDER"]
                c.font = style["NORMAL_FONT"]
                if fill.fill_type:
                    c.fill = fill
                if col_idx in [3, 4]:
                    c.alignment = style["CENTER"]
                elif col_idx == 5:
                    c.alignment = style["RIGHT"]
                    c.number_format = '"$"#,##0.00'
                    
        stock_widths = [30, 20, 15, 18, 15]
        for idx, width in enumerate(stock_widths):
            ws3.column_dimensions[openpyxl.utils.get_column_letter(idx + 1)].width = width
            
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = f"daily_report_{target_date}.xlsx"
        
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
        
    elif format == "pdf":
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        
        title_style, subtitle_style, _, _ = get_pdf_styles()
        elements = []
        
        elements.append(Paragraph(f"Daily Report - {target_date}", title_style))
        elements.append(Spacer(1, 10))
        
        # KPI summary
        kpi_data = {
            "Total Orders": str(report_data["kpis"]["total_orders"]),
            "Total Revenue": f"${report_data['kpis']['total_revenue']}",
            "Books Sold": str(report_data["kpis"]["total_books_sold"]),
            "Net Profit": f"${report_data['kpis']['total_profit']}",
            "Avg Order Value": f"${report_data['kpis']['avg_order_value']}"
        }
        elements.append(make_pdf_kpi_table(kpi_data))
        elements.append(Spacer(1, 15))
        
        # Orders Table
        elements.append(Paragraph("Orders Today", subtitle_style))
        orders_headers = ["ID", "Customer", "Books", "Delivery", "Partner", "Payment", "Total", "Profit", "Status"]
        orders_rows = []
        for o in report_data["orders_table"]:
            orders_rows.append([
                o["order_id"], o["customer"], o["books_count"], o["delivery_way"],
                o["delivery_partner"], o["payment_method"], f"${o['total']}", f"${o['profit']}", o["status"]
            ])
        orders_widths = [60, 110, 50, 80, 70, 70, 70, 70, 70]
        elements.append(make_pdf_table(orders_headers, orders_rows, orders_widths))
        elements.append(Spacer(1, 15))
        
        # Stock Table
        elements.append(Paragraph("Stock Movement Today", subtitle_style))
        stock_headers = ["Book Title", "Category", "Sold Today", "Remaining Stock", "Revenue Today"]
        stock_rows = []
        for s in report_data["stock_table"]:
            stock_rows.append([
                s["book_title"], s["category"], s["stock_out_today"], s["remaining_stock"], f"${s['revenue_today']}"
            ])
        stock_widths = [260, 140, 80, 100, 100]
        elements.append(make_pdf_table(stock_headers, stock_rows, stock_widths))
        
        doc.build(elements)
        buffer.seek(0)
        filename = f"daily_report_{target_date}.pdf"
        
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
        
    else:
        raise HTTPException(status_code=400, detail="Invalid format specified. Must be 'excel' or 'pdf'")

@app.get("/api/v1/admin/reports/weekly/export", tags=["Reports"])
def export_weekly_report(
    date: str | None = Query(None),
    format: str = Query("excel"),
    current_user: TBL_AUTH_USER = Depends(require_admin),
    db: Session = Depends(get_db)
):
    report_data = fetch_weekly_data(date, db)
    week_start = report_data["week_start"]
    week_end = report_data["week_end"]
    
    if format == "excel":
        wb = openpyxl.Workbook()
        style = get_excel_styles()
        
        # --- Sheet 1: Summary ---
        ws1 = wb.active
        ws1.title = "Summary"
        ws1.views.sheetView[0].showGridLines = True
        
        ws1.append(["Weekly Report Summary", f"{week_start} to {week_end}"])
        ws1.cell(1, 1).font = Font(bold=True, size=14, name="Arial")
        ws1.append([])
        
        ws1.append(["KPI Metric", "Value"])
        ws1.cell(3, 1).fill = style["GREEN_FILL"]
        ws1.cell(3, 1).font = style["HEADER_FONT"]
        ws1.cell(3, 2).fill = style["GREEN_FILL"]
        ws1.cell(3, 2).font = style["HEADER_FONT"]
        
        kpi_rows = [
            ("Total Orders", int(report_data["kpis"]["total_orders"])),
            ("Total Revenue", f"${report_data['kpis']['total_revenue']}"),
            ("Total Books Sold", int(report_data["kpis"]["total_books_sold"])),
            ("Net Profit", f"${report_data['kpis']['total_profit']}"),
            ("Avg Daily Revenue", f"${report_data['kpis']['avg_daily_revenue']}")
        ]
        for row in kpi_rows:
            ws1.append(row)
            ws1.cell(ws1.max_row, 1).font = style["BOLD_FONT"]
            ws1.cell(ws1.max_row, 1).border = style["BORDER"]
            ws1.cell(ws1.max_row, 2).border = style["BORDER"]
            
        ws1.column_dimensions['A'].width = 25
        ws1.column_dimensions['B'].width = 20
        
        # --- Sheet 2: Daily Sales ---
        ws2 = wb.create_sheet("Daily Sales")
        ws2.views.sheetView[0].showGridLines = True
        
        headers = ["Date", "Orders", "Revenue", "Books Sold", "Profit"]
        ws2.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws2.cell(1, col_idx)
            cell.fill = style["GREEN_FILL"]
            cell.font = style["HEADER_FONT"]
            cell.alignment = style["CENTER"]
            
        for i, d in enumerate(report_data["daily_table"]):
            row_data = [
                d["date"], int(d["orders"]), float(d["revenue"]),
                int(d["books_sold"]), float(d["profit"])
            ]
            ws2.append(row_data)
            curr_row = ws2.max_row
            fill = style["LIGHT_FILL"] if i % 2 == 0 else PatternFill(fill_type=None)
            
            for col_idx in range(1, len(headers) + 1):
                c = ws2.cell(curr_row, col_idx)
                c.border = style["BORDER"]
                c.font = style["NORMAL_FONT"]
                if fill.fill_type:
                    c.fill = fill
                if col_idx in [1, 2, 4]:
                    c.alignment = style["CENTER"]
                elif col_idx in [3, 5]:
                    c.alignment = style["RIGHT"]
                    c.number_format = '"$"#,##0.00'
                    
        column_widths = [22, 12, 15, 12, 15]
        for idx, width in enumerate(column_widths):
            ws2.column_dimensions[openpyxl.utils.get_column_letter(idx + 1)].width = width
            
        # --- Sheet 3: Stock ---
        ws3 = wb.create_sheet("Stock")
        ws3.views.sheetView[0].showGridLines = True
        
        stock_headers = ["Book Title", "Category", "Opening Stock", "Stock In", "Stock Out", "Closing Stock", "Total Revenue"]
        ws3.append(stock_headers)
        for col_idx in range(1, len(stock_headers) + 1):
            cell = ws3.cell(1, col_idx)
            cell.fill = style["GREEN_FILL"]
            cell.font = style["HEADER_FONT"]
            cell.alignment = style["CENTER"]
            
        for i, s in enumerate(report_data["stock_table"]):
            row_data = [
                s["book_title"], s["category"], int(s["opening_stock"]),
                int(s["stock_in"]), int(s["stock_out"]), int(s["closing_stock"]), float(s["total_revenue"])
            ]
            ws3.append(row_data)
            curr_row = ws3.max_row
            fill = style["LIGHT_FILL"] if i % 2 == 0 else PatternFill(fill_type=None)
            
            for col_idx in range(1, len(stock_headers) + 1):
                c = ws3.cell(curr_row, col_idx)
                c.border = style["BORDER"]
                c.font = style["NORMAL_FONT"]
                if fill.fill_type:
                    c.fill = fill
                if col_idx in [3, 4, 5, 6]:
                    c.alignment = style["CENTER"]
                elif col_idx == 7:
                    c.alignment = style["RIGHT"]
                    c.number_format = '"$"#,##0.00'
                    
        stock_widths = [30, 20, 15, 12, 12, 15, 15]
        for idx, width in enumerate(stock_widths):
            ws3.column_dimensions[openpyxl.utils.get_column_letter(idx + 1)].width = width
            
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = f"weekly_report_{week_start}_{week_end}.xlsx"
        
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
        
    elif format == "pdf":
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        
        title_style, subtitle_style, _, _ = get_pdf_styles()
        elements = []
        
        elements.append(Paragraph(f"Weekly Report ({week_start} to {week_end})", title_style))
        elements.append(Spacer(1, 10))
        
        # KPI summary
        kpi_data = {
            "Total Orders": str(report_data["kpis"]["total_orders"]),
            "Total Revenue": f"${report_data['kpis']['total_revenue']}",
            "Books Sold": str(report_data["kpis"]["total_books_sold"]),
            "Net Profit": f"${report_data['kpis']['total_profit']}",
            "Avg Daily Revenue": f"${report_data['kpis']['avg_daily_revenue']}"
        }
        elements.append(make_pdf_kpi_table(kpi_data, col_width=130))
        elements.append(Spacer(1, 15))
        
        # Daily table
        elements.append(Paragraph("Daily Sales Trend", subtitle_style))
        daily_headers = ["Date", "Orders", "Revenue", "Books Sold", "Profit"]
        daily_rows = []
        for d in report_data["daily_table"]:
            daily_rows.append([
                d["date"], d["orders"], f"${d['revenue']}", d["books_sold"], f"${d['profit']}"
            ])
        daily_widths = [200, 100, 150, 100, 150]
        elements.append(make_pdf_table(daily_headers, daily_rows, daily_widths))
        elements.append(Spacer(1, 15))
        
        # Stock Table
        elements.append(Paragraph("Stock Movement (Weekly)", subtitle_style))
        stock_headers = ["Book Title", "Category", "Opening Stock", "Stock In", "Stock Out", "Closing Stock", "Revenue"]
        stock_rows = []
        for s in report_data["stock_table"]:
            stock_rows.append([
                s["book_title"], s["category"], s["opening_stock"], s["stock_in"],
                s["stock_out"], s["closing_stock"], f"${s['total_revenue']}"
            ])
        stock_widths = [200, 120, 90, 70, 70, 80, 90]
        elements.append(make_pdf_table(stock_headers, stock_rows, stock_widths))
        
        doc.build(elements)
        buffer.seek(0)
        filename = f"weekly_report_{week_start}_{week_end}.pdf"
        
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
        
    else:
        raise HTTPException(status_code=400, detail="Invalid format specified. Must be 'excel' or 'pdf'")

@app.get("/api/v1/admin/reports/monthly/export", tags=["Reports"])
def export_monthly_report(
    month: int | None = Query(None),
    year: int | None = Query(None),
    format: str = Query("excel"),
    current_user: TBL_AUTH_USER = Depends(require_admin),
    db: Session = Depends(get_db)
):
    report_data = fetch_monthly_data(month, year, db)
    month_name = report_data["month_name"].replace(" ", "_")
    
    if format == "excel":
        wb = openpyxl.Workbook()
        style = get_excel_styles()
        
        # --- Sheet 1: Summary ---
        ws1 = wb.active
        ws1.title = "Summary"
        ws1.views.sheetView[0].showGridLines = True
        
        ws1.append(["Monthly Report Summary", report_data["month_name"]])
        ws1.cell(1, 1).font = Font(bold=True, size=14, name="Arial")
        ws1.append([])
        
        ws1.append(["KPI Metric", "Value"])
        ws1.cell(3, 1).fill = style["GREEN_FILL"]
        ws1.cell(3, 1).font = style["HEADER_FONT"]
        ws1.cell(3, 2).fill = style["GREEN_FILL"]
        ws1.cell(3, 2).font = style["HEADER_FONT"]
        
        kpi_rows = [
            ("Total Orders", int(report_data["kpis"]["total_orders"])),
            ("Total Revenue", f"${report_data['kpis']['total_revenue']}"),
            ("Total Books Sold", int(report_data["kpis"]["total_books_sold"])),
            ("Net Profit", f"${report_data['kpis']['total_profit']}"),
            ("Best Selling Book", report_data["kpis"]["best_selling_book"]),
            ("Avg Daily Revenue", f"${report_data['kpis']['avg_daily_revenue']}")
        ]
        for row in kpi_rows:
            ws1.append(row)
            ws1.cell(ws1.max_row, 1).font = style["BOLD_FONT"]
            ws1.cell(ws1.max_row, 1).border = style["BORDER"]
            ws1.cell(ws1.max_row, 2).border = style["BORDER"]
            
        ws1.column_dimensions['A'].width = 25
        ws1.column_dimensions['B'].width = 30
        
        # --- Sheet 2: Daily Sales ---
        ws2 = wb.create_sheet("Daily Sales")
        ws2.views.sheetView[0].showGridLines = True
        
        headers = ["Date", "Orders", "Revenue", "Books Sold", "Profit"]
        ws2.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws2.cell(1, col_idx)
            cell.fill = style["GREEN_FILL"]
            cell.font = style["HEADER_FONT"]
            cell.alignment = style["CENTER"]
            
        for i, d in enumerate(report_data["daily_table"]):
            row_data = [
                d["date"], int(d["orders"]), float(d["revenue"]),
                int(d["books_sold"]), float(d["profit"])
            ]
            ws2.append(row_data)
            curr_row = ws2.max_row
            fill = style["LIGHT_FILL"] if i % 2 == 0 else PatternFill(fill_type=None)
            
            for col_idx in range(1, len(headers) + 1):
                c = ws2.cell(curr_row, col_idx)
                c.border = style["BORDER"]
                c.font = style["NORMAL_FONT"]
                if fill.fill_type:
                    c.fill = fill
                if col_idx in [1, 2, 4]:
                    c.alignment = style["CENTER"]
                elif col_idx in [3, 5]:
                    c.alignment = style["RIGHT"]
                    c.number_format = '"$"#,##0.00'
                    
        column_widths = [18, 12, 15, 12, 15]
        for idx, width in enumerate(column_widths):
            ws2.column_dimensions[openpyxl.utils.get_column_letter(idx + 1)].width = width
            
        # --- Sheet 3: Stock ---
        ws3 = wb.create_sheet("Stock")
        ws3.views.sheetView[0].showGridLines = True
        
        stock_headers = ["Book Title", "Category", "Opening Stock", "Stock In", "Stock Out", "Closing Stock", "Total Revenue"]
        ws3.append(stock_headers)
        for col_idx in range(1, len(stock_headers) + 1):
            cell = ws3.cell(1, col_idx)
            cell.fill = style["GREEN_FILL"]
            cell.font = style["HEADER_FONT"]
            cell.alignment = style["CENTER"]
            
        for i, s in enumerate(report_data["stock_table"]):
            row_data = [
                s["book_title"], s["category"], int(s["opening_stock"]),
                int(s["stock_in"]), int(s["stock_out"]), int(s["closing_stock"]), float(s["total_revenue"])
            ]
            ws3.append(row_data)
            curr_row = ws3.max_row
            fill = style["LIGHT_FILL"] if i % 2 == 0 else PatternFill(fill_type=None)
            
            for col_idx in range(1, len(stock_headers) + 1):
                c = ws3.cell(curr_row, col_idx)
                c.border = style["BORDER"]
                c.font = style["NORMAL_FONT"]
                if fill.fill_type:
                    c.fill = fill
                if col_idx in [3, 4, 5, 6]:
                    c.alignment = style["CENTER"]
                elif col_idx == 7:
                    c.alignment = style["RIGHT"]
                    c.number_format = '"$"#,##0.00'
                    
        stock_widths = [30, 20, 15, 12, 12, 15, 15]
        for idx, width in enumerate(stock_widths):
            ws3.column_dimensions[openpyxl.utils.get_column_letter(idx + 1)].width = width
            
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = f"monthly_report_{month_name}.xlsx"
        
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
        
    elif format == "pdf":
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        
        title_style, subtitle_style, _, _ = get_pdf_styles()
        elements = []
        
        elements.append(Paragraph(f"Monthly Report - {report_data['month_name']}", title_style))
        elements.append(Spacer(1, 10))
        
        # KPI summary
        kpi_data = {
            "Total Orders": str(report_data["kpis"]["total_orders"]),
            "Total Revenue": f"${report_data['kpis']['total_revenue']}",
            "Books Sold": str(report_data["kpis"]["total_books_sold"]),
            "Net Profit": f"${report_data['kpis']['total_profit']}",
            "Best Selling": str(report_data["kpis"]["best_selling_book"]),
            "Avg Daily Revenue": f"${report_data['kpis']['avg_daily_revenue']}"
        }
        elements.append(make_pdf_kpi_table(kpi_data, col_width=115))
        elements.append(Spacer(1, 15))
        
        # Daily table
        elements.append(Paragraph("Daily Sales Trend", subtitle_style))
        daily_headers = ["Date", "Orders", "Revenue", "Books Sold", "Profit"]
        daily_rows = []
        for d in report_data["daily_table"]:
            daily_rows.append([
                d["date"], d["orders"], f"${d['revenue']}", d["books_sold"], f"${d['profit']}"
            ])
        daily_widths = [180, 100, 150, 120, 150]
        elements.append(make_pdf_table(daily_headers, daily_rows, daily_widths))
        elements.append(Spacer(1, 15))
        
        # Stock Table
        elements.append(Paragraph("Stock Movement", subtitle_style))
        stock_headers = ["Book Title", "Category", "Opening Stock", "Stock In", "Stock Out", "Closing Stock", "Revenue"]
        stock_rows = []
        for s in report_data["stock_table"]:
            stock_rows.append([
                s["book_title"], s["category"], s["opening_stock"], s["stock_in"],
                s["stock_out"], s["closing_stock"], f"${s['total_revenue']}"
            ])
        stock_widths = [200, 120, 90, 70, 70, 80, 90]
        elements.append(make_pdf_table(stock_headers, stock_rows, stock_widths))
        
        doc.build(elements)
        buffer.seek(0)
        filename = f"monthly_report_{month_name}.pdf"
        
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
        
    else:
        raise HTTPException(status_code=400, detail="Invalid format specified. Must be 'excel' or 'pdf'")

@app.get("/api/v1/admin/reports/yearly/export", tags=["Reports"])
def export_yearly_report(
    year: int | None = Query(None),
    format: str = Query("excel"),
    current_user: TBL_AUTH_USER = Depends(require_admin),
    db: Session = Depends(get_db)
):
    report_data = fetch_yearly_data(year, db)
    target_year = report_data["year"]
    
    if format == "excel":
        wb = openpyxl.Workbook()
        style = get_excel_styles()
        
        # --- Sheet 1: Summary ---
        ws1 = wb.active
        ws1.title = "Summary"
        ws1.views.sheetView[0].showGridLines = True
        
        ws1.append(["Yearly Report Summary", str(target_year)])
        ws1.cell(1, 1).font = Font(bold=True, size=14, name="Arial")
        ws1.append([])
        
        ws1.append(["KPI Metric", "Value"])
        ws1.cell(3, 1).fill = style["GREEN_FILL"]
        ws1.cell(3, 1).font = style["HEADER_FONT"]
        ws1.cell(3, 2).fill = style["GREEN_FILL"]
        ws1.cell(3, 2).font = style["HEADER_FONT"]
        
        kpi_rows = [
            ("Total Orders", int(report_data["kpis"]["total_orders"])),
            ("Total Revenue", f"${report_data['kpis']['total_revenue']}"),
            ("Total Books Sold", int(report_data["kpis"]["total_books_sold"])),
            ("Net Profit", f"${report_data['kpis']['total_profit']}"),
            ("Average Monthly Revenue", f"${report_data['kpis']['avg_monthly_revenue']}"),
            ("Best Month", report_data["kpis"]["best_month"])
        ]
        for row in kpi_rows:
            ws1.append(row)
            ws1.cell(ws1.max_row, 1).font = style["BOLD_FONT"]
            ws1.cell(ws1.max_row, 1).border = style["BORDER"]
            ws1.cell(ws1.max_row, 2).border = style["BORDER"]
            
        ws1.column_dimensions['A'].width = 25
        ws1.column_dimensions['B'].width = 20
        
        # --- Sheet 2: Monthly Sales ---
        ws2 = wb.create_sheet("Monthly Sales")
        ws2.views.sheetView[0].showGridLines = True
        
        headers = ["Month", "Orders", "Revenue", "Books Sold", "Profit"]
        ws2.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws2.cell(1, col_idx)
            cell.fill = style["GREEN_FILL"]
            cell.font = style["HEADER_FONT"]
            cell.alignment = style["CENTER"]
            
        for i, m in enumerate(report_data["monthly_table"]):
            row_data = [
                m["month"], int(m["orders"]), float(m["revenue"]),
                int(m["books_sold"]), float(m["profit"])
            ]
            ws2.append(row_data)
            curr_row = ws2.max_row
            fill = style["LIGHT_FILL"] if i % 2 == 0 else PatternFill(fill_type=None)
            
            for col_idx in range(1, len(headers) + 1):
                c = ws2.cell(curr_row, col_idx)
                c.border = style["BORDER"]
                c.font = style["NORMAL_FONT"]
                if fill.fill_type:
                    c.fill = fill
                if col_idx in [1, 2, 4]:
                    c.alignment = style["CENTER"]
                elif col_idx in [3, 5]:
                    c.alignment = style["RIGHT"]
                    c.number_format = '"$"#,##0.00'
                    
        column_widths = [18, 12, 15, 12, 15]
        for idx, width in enumerate(column_widths):
            ws2.column_dimensions[openpyxl.utils.get_column_letter(idx + 1)].width = width
            
        # --- Sheet 3: Stock ---
        ws3 = wb.create_sheet("Stock")
        ws3.views.sheetView[0].showGridLines = True
        
        stock_headers = ["Book Title", "Total Sold", "Total Revenue", "Remaining Stock"]
        ws3.append(stock_headers)
        for col_idx in range(1, len(stock_headers) + 1):
            cell = ws3.cell(1, col_idx)
            cell.fill = style["GREEN_FILL"]
            cell.font = style["HEADER_FONT"]
            cell.alignment = style["CENTER"]
            
        for i, s in enumerate(report_data["stock_table"]):
            row_data = [
                s["book_title"], int(s["total_sold"]), float(s["total_revenue"]), int(s["remaining"])
            ]
            ws3.append(row_data)
            curr_row = ws3.max_row
            fill = style["LIGHT_FILL"] if i % 2 == 0 else PatternFill(fill_type=None)
            
            for col_idx in range(1, len(stock_headers) + 1):
                c = ws3.cell(curr_row, col_idx)
                c.border = style["BORDER"]
                c.font = style["NORMAL_FONT"]
                if fill.fill_type:
                    c.fill = fill
                if col_idx in [2, 4]:
                    c.alignment = style["CENTER"]
                elif col_idx == 3:
                    c.alignment = style["RIGHT"]
                    c.number_format = '"$"#,##0.00'
                    
        stock_widths = [35, 15, 18, 18]
        for idx, width in enumerate(stock_widths):
            ws3.column_dimensions[openpyxl.utils.get_column_letter(idx + 1)].width = width
            
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = f"yearly_report_{target_year}.xlsx"
        
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
        
    elif format == "pdf":
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        
        title_style, subtitle_style, _, _ = get_pdf_styles()
        elements = []
        
        elements.append(Paragraph(f"Yearly Report - {target_year}", title_style))
        elements.append(Spacer(1, 10))
        
        # KPI summary
        kpi_data = {
            "Total Orders": str(report_data["kpis"]["total_orders"]),
            "Total Revenue": f"${report_data['kpis']['total_revenue']}",
            "Books Sold": str(report_data["kpis"]["total_books_sold"]),
            "Net Profit": f"${report_data['kpis']['total_profit']}",
            "Avg Monthly Revenue": f"${report_data['kpis']['avg_monthly_revenue']}",
            "Best Month": str(report_data["kpis"]["best_month"])
        }
        elements.append(make_pdf_kpi_table(kpi_data, col_width=115))
        elements.append(Spacer(1, 15))
        
        # Monthly table
        elements.append(Paragraph("Monthly Performance Summary", subtitle_style))
        monthly_headers = ["Month", "Orders", "Revenue", "Books Sold", "Profit"]
        monthly_rows = []
        for m in report_data["monthly_table"]:
            monthly_rows.append([
                m["month"], m["orders"], f"${m['revenue']}", m["books_sold"], f"${m['profit']}"
            ])
        monthly_widths = [180, 100, 150, 120, 150]
        elements.append(make_pdf_table(monthly_headers, monthly_rows, monthly_widths))
        elements.append(Spacer(1, 15))
        
        # Stock Table
        elements.append(Paragraph("Stock Summary", subtitle_style))
        stock_headers = ["Book Title", "Total Sold", "Total Revenue", "Remaining Stock"]
        stock_rows = []
        for s in report_data["stock_table"]:
            stock_rows.append([
                s["book_title"], s["total_sold"], f"${s['total_revenue']}", s["remaining"]
            ])
        stock_widths = [300, 140, 140, 140]
        elements.append(make_pdf_table(stock_headers, stock_rows, stock_widths))
        
        doc.build(elements)
        buffer.seek(0)
        filename = f"yearly_report_{target_year}.pdf"
        
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
        
    else:
        raise HTTPException(status_code=400, detail="Invalid format specified. Must be 'excel' or 'pdf'")
