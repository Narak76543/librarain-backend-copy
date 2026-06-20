import logging
from fastapi import Depends, File, Query, UploadFile, status
from fastapi.concurrency import run_in_threadpool
from fastapi.encoders import jsonable_encoder
from sqlalchemy.orm import Session, joinedload

from config import configs
from core.db import get_db
from main import app
from api.auth_user.views import response
from api.auth_user.models import TBL_AUTH_USER
from api.auth_user.security import require_admin
from api.books.models import TBL_BOOK, TBL_STOCK_HISTORY, TBL_STOCK_IN
from api.orders.models import TBL_ORDER, TBL_ORDER_ITEM
from sqlalchemy import func, or_
from api.books import schemas
from api.categories.models import TBL_CATEGORY
from datetime import datetime, timedelta, timezone
from fastapi.responses import StreamingResponse
import csv
import io

logger = logging.getLogger(__name__)


# ========= helpers =============================================
def upload_cover_to_cloudinary(contents: bytes, book_id: str) -> str:
    import cloudinary
    import cloudinary.uploader

    cloudinary.config(
        cloud_name = configs.CLOUDINARY_CLOUD_NAME,
        api_key    = configs.CLOUDINARY_API_KEY,
        api_secret = configs.CLOUDINARY_API_SECRET,
    )
    result = cloudinary.uploader.upload(
        contents,
        folder         = f"books/{book_id}",
        resource_type  = "image",
        transformation = [{"width": 400, "height": 560, "crop": "fill"}],
    )
    return result["secure_url"]



# ================= GET /books ================================
@app.get("/api/v1/books", tags=["Books"])
def get_books(
    search   : str | None   = Query(None),
    category : str | None   = Query(None),
    min_price: float | None = Query(None),
    max_price: float | None = Query(None),
    featured : bool | None  = Query(None),
    sort     : str | None   = Query(None),
    limit    : int          = Query(10, ge=1, le=100),
    offset   : int          = Query(0,  ge=0),
    db       : Session      = Depends(get_db),
):
    query = (
        db.query(TBL_BOOK)
        .filter(TBL_BOOK.is_active.is_(True))
    )

    if search:
        search_terms = search.split()
        conditions = []
        for term in search_terms:
            if len(term) > 2:  # Ignore small words like 'A', 'of', 'to'
                conditions.append(TBL_BOOK.title.ilike(f"%{term}%"))
                conditions.append(TBL_BOOK.author.ilike(f"%{term}%"))
        
        if conditions:
            query = query.filter(or_(*conditions))
        else:
            # Fallback if the search only contained short words
            query = query.filter(
                or_(
                    TBL_BOOK.title.ilike(f"%{search}%"),
                    TBL_BOOK.author.ilike(f"%{search}%"),
                )
            )

    if category:
        query = query.join(TBL_CATEGORY).filter(
            TBL_CATEGORY.slug == category
        )

    if min_price is not None:
        query = query.filter(TBL_BOOK.price >= min_price)

    if max_price is not None:
        query = query.filter(TBL_BOOK.price <= max_price)

    if featured is True:
        query = query.filter(TBL_BOOK.featured.is_(True))

    if sort == "newest":
        query = query.order_by(TBL_BOOK.created_at.desc())
    elif sort == "price_asc":
        query = query.order_by(TBL_BOOK.price.asc())
    elif sort == "price_desc":
        query = query.order_by(TBL_BOOK.price.desc())
    elif sort == "rating":
        query = query.order_by(TBL_BOOK.rating_average.desc())
    else:
        query = query.order_by(TBL_BOOK.created_at.desc())

    total = query.count()

    books = (
        query
        .offset(offset)
        .limit(limit)
        .all()
    )

    data = jsonable_encoder([
        schemas.BookResponse.model_validate(b) for b in books
    ])

    return response(
        ok          = True,
        status_code = status.HTTP_200_OK,
        message     = "Books retrieved successfully",
        data        = {
            "total" : total,
            "limit" : limit,
            "offset": offset,
            "books" : data,
        },
    )


# ================= GET /books/{id} ==========================
@app.get("/api/v1/books/{book_id}", tags=["Books"])
def get_book(
    book_id: str,
    db     : Session = Depends(get_db),
):
    book = (
        db.query(TBL_BOOK)
        .filter(
            TBL_BOOK.id       == book_id,
            TBL_BOOK.is_active.is_(True),
        )
        .first()
    )

    if not book:
        return response(
            ok          = False,
            status_code = status.HTTP_404_NOT_FOUND,
            message     = "Book not found",
        )

    return response(
        ok          = True,
        status_code = status.HTTP_200_OK,
        message     = "Book retrieved successfully",
        data        = jsonable_encoder(
            schemas.BookResponse.model_validate(book)
        ),
    )


# ================= POST /books ==============================
@app.post("/api/v1/books", tags=["Books"])
def create_book(
    payload     : schemas.BookCreate,
    current_user: TBL_AUTH_USER = Depends(require_admin),
    db          : Session       = Depends(get_db),
):
    book = TBL_BOOK(**payload.model_dump())
    db.add(book)
    db.commit()
    db.refresh(book)

    if book.stock > 0:
        history = TBL_STOCK_HISTORY(
            book_id    = book.id,
            quantity   = book.stock,
            cost_price = book.cost_price,
            sale_price = book.price,
            user_id    = current_user.id,
        )
        db.add(history)
        db.commit()

    return response(
        ok          = True,
        status_code = status.HTTP_201_CREATED,
        message     = "Book created successfully",
        data        = jsonable_encoder(
            schemas.BookResponse.model_validate(book)
        ),
    )


# ================= POST /books/{id}/cover ===================
@app.post("/api/v1/books/{book_id}/cover", tags=["Books"])
async def upload_book_cover(
    book_id     : str,
    file        : UploadFile    = File(...),
    current_user: TBL_AUTH_USER = Depends(require_admin),
    db          : Session       = Depends(get_db),
):
    allowed_types = {"image/jpeg", "image/png", "image/webp"}

    if file.content_type not in allowed_types:
        return response(
            ok          = False,
            status_code = status.HTTP_400_BAD_REQUEST,
            message     = "Only JPEG, PNG and WebP images are allowed",
        )

    book = (
        db.query(TBL_BOOK)
        .filter(TBL_BOOK.id == book_id)
        .first()
    )

    if not book:
        return response(
            ok          = False,
            status_code = status.HTTP_404_NOT_FOUND,
            message     = "Book not found",
        )

    try:
        contents  = await file.read()
        MAX_FILE_SIZE = 5 * 1024 * 1024  # 5MB
        if len(contents) > MAX_FILE_SIZE:
            return response(
                ok          = False,
                status_code = status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
                message     = "File too large. Maximum size is 5MB",
            )
        cover_url = await run_in_threadpool(upload_cover_to_cloudinary, contents, book_id)
    except Exception:
        logger.exception("Failed to upload book cover to Cloudinary")
        return response(
            ok          = False,
            status_code = status.HTTP_502_BAD_GATEWAY,
            message     = "Failed to upload cover. Check Cloudinary configuration.",
        )

    book.cover_url = cover_url
    db.commit()
    db.refresh(book)

    return response(
        ok          = True,
        status_code = status.HTTP_200_OK,
        message     = "Book cover uploaded successfully",
        data        = jsonable_encoder(
            schemas.BookResponse.model_validate(book)
        ),
    )


# ================= POST /books/{id}/stock-in =================
@app.post("/api/v1/books/{book_id}/stock-in", tags=["Books"])
def stock_in_book(
    book_id     : str,
    payload     : schemas.StockInCreate,
    current_user: TBL_AUTH_USER = Depends(require_admin),
    db          : Session       = Depends(get_db),
):
    book = (
        db.query(TBL_BOOK)
        .filter(TBL_BOOK.id == book_id)
        .first()
    )

    if not book:
        return response(
            ok          = False,
            status_code = status.HTTP_404_NOT_FOUND,
            message     = "Book not found",
        )

    # 1. Update book
    book.stock      += payload.quantity
    book.cost_price  = payload.cost_price
    book.price       = payload.sale_price
    
    # 2. Log to history
    history = TBL_STOCK_HISTORY(
        book_id    = book.id,
        quantity   = payload.quantity,
        cost_price = payload.cost_price,
        sale_price = payload.sale_price,
        user_id    = current_user.id,
    )
    db.add(history)
    
    # 3. Log to tbl_stock_in for the daily report
    stock_in = TBL_STOCK_IN(
        book_id    = book.id,
        quantity   = payload.quantity,
        cost_price = payload.cost_price,
        total_cost = payload.cost_price * payload.quantity,
        note       = payload.notes,
        created_by = current_user.id,
    )
    db.add(stock_in)
    
    db.commit()
    db.refresh(book)

    return response(
        ok          = True,
        status_code = status.HTTP_200_OK,
        message     = "Stock added successfully",
        data        = jsonable_encoder(
            schemas.BookResponse.model_validate(book)
        ),
    )


# ================= PUT /books/{id} ==========================
@app.put("/api/v1/books/{book_id}", tags=["Books"])
def update_book(
    book_id     : str,
    payload     : schemas.BookUpdate,
    current_user: TBL_AUTH_USER = Depends(require_admin),
    db          : Session       = Depends(get_db),
):
    book = (
        db.query(TBL_BOOK)
        .filter(TBL_BOOK.id == book_id)
        .first()
    )

    if not book:
        return response(
            ok          = False,
            status_code = status.HTTP_404_NOT_FOUND,
            message     = "Book not found",
        )

    update_data = payload.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(book, field, value)

    db.commit()
    db.refresh(book)

    return response(
        ok          = True,
        status_code = status.HTTP_200_OK,
        message     = "Book updated successfully",
        data        = jsonable_encoder(
            schemas.BookResponse.model_validate(book)
        ),
    )


# ================= DELETE /books/{id} =======================
@app.delete("/api/v1/books/{book_id}", tags=["Books"])
def delete_book(
    book_id     : str,
    current_user: TBL_AUTH_USER = Depends(require_admin),
    db          : Session       = Depends(get_db),
):
    book = (
        db.query(TBL_BOOK)
        .filter(TBL_BOOK.id == book_id)
        .first()
    )

    if not book:
        return response(
            ok          = False,
            status_code = status.HTTP_404_NOT_FOUND,
            message     = "Book not found",
        )

    book.is_active = False
    db.commit()

    return response(
        ok          = True,
        status_code = status.HTTP_200_OK,
        message     = "Book deleted successfully",
    )



@app.get("/api/v1/admin/reports/stock", tags=["Reports"])
def stock_report(
    period    : str | None = Query(None),  # daily/weekly/monthly/yearly
    date_from : str | None = Query(None),  # custom range
    date_to   : str | None = Query(None),
    export    : str | None = Query(None),  # csv/excel
    current_user: TBL_AUTH_USER = Depends(require_admin),
    db        : Session         = Depends(get_db),
):
    now = datetime.now(timezone.utc)

    # ====== Date range logic ============
    if period == "daily":
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        end   = now

    elif period == "weekly":
        start = now - timedelta(days=now.weekday())
        start = start.replace(hour=0, minute=0, second=0, microsecond=0)
        end   = now

    elif period == "monthly":
        start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        end   = now

    elif period == "yearly":
        start = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        end   = now

    elif date_from and date_to:
        start = datetime.fromisoformat(date_from).replace(tzinfo=timezone.utc)
        end   = datetime.fromisoformat(date_to  ).replace(tzinfo=timezone.utc)

    else:
        # Default — all time
        start = None
        end   = None

    # ================= Query books ================================
    query = (
        db.query(TBL_BOOK)
        .filter(TBL_BOOK.is_active.is_(True))
    )

    if start and end:
        query = query.filter(
            TBL_BOOK.created_at >= start,
            TBL_BOOK.created_at <= end,
        )

    books = query.order_by(TBL_BOOK.created_at.desc()).all()

    # ================ Calculate KPIs & Stock-In Data ======================
    totals = db.query(
        func.sum(TBL_STOCK_HISTORY.quantity).label('qty'),
        func.sum(TBL_STOCK_HISTORY.cost_price * TBL_STOCK_HISTORY.quantity).label('cost'),
        func.sum(TBL_STOCK_HISTORY.sale_price * TBL_STOCK_HISTORY.quantity).label('revenue')
    ).first()
    
    total_books_added = int(totals.qty or 0)
    total_cost_invested = float(totals.cost or 0.0)
    total_potential_revenue = float(totals.revenue or 0.0)
    total_potential_profit = total_potential_revenue - total_cost_invested

    stock_in_by_book_query = db.query(
        TBL_STOCK_HISTORY.book_id,
        func.sum(TBL_STOCK_HISTORY.quantity).label('qty'),
        func.sum(TBL_STOCK_HISTORY.cost_price * TBL_STOCK_HISTORY.quantity).label('cost')
    ).group_by(TBL_STOCK_HISTORY.book_id).all()
    
    stock_in_by_book = {
        row.book_id: {"qty": int(row.qty or 0), "cost": float(row.cost or 0.0)}
        for row in stock_in_by_book_query
    }

    # ================ Serialize ===========================================
    rows = []
    total_current_stock = 0
    for b in books:
        total_current_stock += b.stock
        cost_price  = float(b.cost_price or 0.0)
        sale_price  = float(b.price)
        stock_value = cost_price * b.stock
        si = stock_in_by_book.get(b.id, {"qty": 0, "cost": 0.0})

        rows.append({
            "book_title":     b.title,
            "author":         b.author,
            "category":       b.category.name if b.category else "—",
            "total_stock_in": si["qty"],
            "total_invested": round(si["cost"], 2),
            "stock":          b.stock,
            "cost_price":     round(cost_price, 2),
            "sale_price":     round(sale_price, 2),
            "stock_value":    round(stock_value, 2),
            "created_at":     b.created_at.strftime("%Y-%m-%d") if b.created_at else "—",
        })

    summary = {
        "total_books":   total_books_added,
        "total_stock":   total_current_stock,
        "total_cost":    round(total_cost_invested,    2),
        "total_revenue": round(total_potential_revenue, 2),
        "profit_margin": round(total_potential_profit, 2),
    }

    # ============= Export CSV ========================
    if export == "csv":
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=[
            "book_title", "author", "category",
            "total_stock_in", "total_invested",
            "stock", "cost_price", "sale_price",
            "stock_value", "created_at",
        ])
        writer.writeheader()
        writer.writerows(rows)

        # Summary rows
        output.write("\n")
        output.write(f"Total Books,{summary['total_books']}\n")
        output.write(f"Total Stock,{summary['total_stock']}\n")
        output.write(f"Total Cost,${ summary['total_cost']}\n")
        output.write(f"Total Revenue,${summary['total_revenue']}\n")
        output.write(f"Profit Margin,${summary['profit_margin']}\n")

        output.seek(0)
        filename = f"stock_report_{now.strftime('%Y%m%d_%H%M%S')}.csv"
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode("utf-8-sig")),
            media_type = "text/csv",
            headers    = {"Content-Disposition": f"attachment; filename={filename}"},
        )

    # =================== Export Excel ======================================
    if export == "excel":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            return response(
                ok          = False,
                status_code = status.HTTP_500_INTERNAL_SERVER_ERROR,
                message     = "openpyxl not installed. Run: pip install openpyxl",
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stock Report"

        # ============ Style helpers ==========================
        GREEN_FILL  = PatternFill("solid", fgColor="059669")
        LIGHT_FILL  = PatternFill("solid", fgColor="ECFDF5")
        HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        BOLD_FONT   = Font(bold=True, name="Arial", size=10)
        NORMAL_FONT = Font(name="Arial", size=10)
        CENTER      = Alignment(horizontal="center", vertical="center")
        LEFT        = Alignment(horizontal="left",   vertical="center")
        thin        = Side(style="thin", color="E5E7EB")
        BORDER      = Border(left=thin, right=thin, top=thin, bottom=thin)

        # ============= Title =======================================
        ws.merge_cells("A1:H1")
        ws["A1"] = "📚 Librarain — Stock Adjustment Report"
        ws["A1"].font      = Font(bold=True, size=14, name="Arial", color="059669")
        ws["A1"].alignment = CENTER
        ws.row_dimensions[1].height = 30

        # ========== Period row ======================
        ws.merge_cells("A2:H2")
        period_label = period.capitalize() if period else "All Time"
        if date_from and date_to:
            period_label = f"{date_from} to {date_to}"
        ws["A2"] = f"Period: {period_label}  |  Generated: {now.strftime('%B %d, %Y %H:%M')}"
        ws["A2"].font      = Font(size=10, name="Arial", color="6B7280")
        ws["A2"].alignment = CENTER
        ws.row_dimensions[2].height = 20

        ws.append([])  # spacer

        # ── Headers ──
        headers = ["#", "Book Title", "Author", "Category",
                   "Total Stock In", "Total Invested",
                   "Current Stock", "Cost Price", "Sale Price",
                   "Stock Value", "Date Added"]
        ws.append(headers)
        header_row = ws.max_row
        for col, header in enumerate(headers, 1):
            cell            = ws.cell(header_row, col)
            cell.value      = header
            cell.font       = HEADER_FONT
            cell.fill       = GREEN_FILL
            cell.alignment  = CENTER
            cell.border     = BORDER
        ws.row_dimensions[header_row].height = 25

        # ── Data rows ──
        for i, row in enumerate(rows, 1):
            fill = PatternFill("solid", fgColor="F9FAFB") if i % 2 == 0 else None
            data = [
                i,
                row["book_title"],
                row["author"],
                row["category"],
                row["total_stock_in"],
                row["total_invested"],
                row["stock"],
                row["cost_price"],
                row["sale_price"],
                row["stock_value"],
                row["created_at"],
            ]
            ws.append(data)
            current_row = ws.max_row
            ws.row_dimensions[current_row].height = 20
            for col in range(1, len(data) + 1):
                cell           = ws.cell(current_row, col)
                cell.font      = NORMAL_FONT
                cell.border    = BORDER
                cell.alignment = CENTER if col in [1, 5, 7, 11] else LEFT
                if fill:
                    cell.fill  = fill
            
            # Formats
            ws.cell(current_row, 6).number_format = '"$"#,##0.00'
            ws.cell(current_row, 8).number_format = '"$"#,##0.00'
            ws.cell(current_row, 9).number_format = '"$"#,##0.00'
            ws.cell(current_row, 10).number_format = '"$"#,##0.00'

        ws.append([])  # spacer

        # ── Summary rows ──
        summary_data = [
            ("Total Books",   summary["total_books"],   None),
            ("Total Stock",   summary["total_stock"],   None),
            ("Total Cost",    f"${summary['total_cost']:.2f}",    "FEF2F2"),
            ("Total Revenue", f"${summary['total_revenue']:.2f}", "ECFDF5"),
            ("Profit Margin", f"${summary['profit_margin']:.2f}", "EFF6FF"),
        ]
        for label, value, bg in summary_data:
            ws.append(["", "", "", "", "", "", label, value, ""])
            r = ws.max_row
            ws.cell(r, 7).font      = BOLD_FONT
            ws.cell(r, 7).alignment = LEFT
            ws.cell(r, 8).font      = BOLD_FONT
            ws.cell(r, 8).alignment = LEFT
            if bg:
                ws.cell(r, 7).fill = PatternFill("solid", fgColor=bg)
                ws.cell(r, 8).fill = PatternFill("solid", fgColor=bg)
            ws.row_dimensions[r].height = 20

        # ── Column widths ──
        col_widths = [5, 35, 20, 15, 8, 12, 12, 14, 14]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(i)
            ].width = w

        # ── Save to buffer ──
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = f"stock_report_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"

        return StreamingResponse(
            buffer,
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers    = {"Content-Disposition": f"attachment; filename={filename}"},
        )

    # ======== JSON response (for web table) ===================
    return response(
        ok          = True,
        status_code = status.HTTP_200_OK,
        message     = "Stock report retrieved successfully",
        data        = {
            "period":  period_label if 'period_label' in dir() else "All Time",
            "rows":    rows,
            "summary": summary,
        },
    )


# ================= GET /reports/sales ==========================
@app.get("/api/v1/admin/reports/sales", tags=["Reports"])
def sales_report(
    limit       : int = Query(10, ge=1, le=100),
    offset      : int = Query(0, ge=0),
    db: Session = Depends(get_db),
    current_user: TBL_AUTH_USER = Depends(require_admin),
):
    order_items = (
        db.query(TBL_ORDER_ITEM)
        .options(joinedload(TBL_ORDER_ITEM.order).joinedload(TBL_ORDER.user), joinedload(TBL_ORDER_ITEM.book))
        .join(TBL_ORDER)
        .filter(func.lower(TBL_ORDER.status) == "completed")
        .order_by(TBL_ORDER.created_at.desc())
        .offset(offset)
        .limit(limit)
        .all()
    )

    rows = []
    total_sales = 0
    total_profit = 0

    for item in order_items:
        sale_p = float(item.price_at_purchase or 0.0) * item.quantity
        cost_p = float(item.cost_price_at_purchase or 0.0) * item.quantity
        profit = sale_p - cost_p

        total_sales += sale_p
        total_profit += profit

        rows.append({
            "order_id": str(item.order_id),
            "date": item.order.created_at.isoformat() if item.order.created_at else None,
            "customer_name": item.order.user.full_name if item.order.user else "Unknown",
            "book_title": item.book.title if item.book else "Deleted Book",
            "quantity": item.quantity,
            "sale_price": float(item.price_at_purchase or 0.0),
            "cost_price": float(item.cost_price_at_purchase or 0.0),
            "profit": round(profit, 2)
        })

    return response(
        ok=True,
        status_code=status.HTTP_200_OK,
        message="Sales report retrieved",
        data={
            "rows": rows,
            "summary": {
                "total_sales": round(total_sales, 2),
                "total_profit": round(total_profit, 2)
            }
        }
    )


# ================= GET /reports/stock-in ==========================
@app.get("/api/v1/admin/reports/stock-in", tags=["Reports"])
def stock_in_report(
    limit       : int = Query(10, ge=1, le=100),
    offset      : int = Query(0, ge=0),
    db: Session = Depends(get_db),
    current_user: TBL_AUTH_USER = Depends(require_admin),
):
    history = (
        db.query(TBL_STOCK_IN)
        .options(joinedload(TBL_STOCK_IN.book), joinedload(TBL_STOCK_IN.user))
        .order_by(TBL_STOCK_IN.created_at.desc())
        .offset(offset)
        .limit(limit)
        .all()
    )

    rows = []
    total_books_added = 0
    total_cost_invested = 0

    for h in history:
        cost = float(h.total_cost or 0.0)
        total_books_added += h.quantity
        total_cost_invested += cost

        rows.append({
            "date": h.created_at.isoformat(),
            "book_title": h.book.title if h.book else "Deleted Book",
            "quantity": h.quantity,
            "cost_price": float(h.cost_price or 0.0),
            "total_cost": float(h.total_cost or 0.0),
            "note": h.note or "",
            "admin": h.user.full_name if h.user else "Unknown",
            "current_stock": h.book.stock if h.book else 0,
        })

    return response(
        ok=True,
        status_code=status.HTTP_200_OK,
        message="Stock in history retrieved",
        data={
            "rows": rows,
            "summary": {
                "total_books_added": total_books_added,
                "total_cost_invested": round(total_cost_invested, 2)
            }
        }
    )


# ================= GET /reports/sales-by-book ==========================
@app.get("/api/v1/admin/reports/sales-by-book", tags=["Reports"])
def sales_by_book_report(
    limit       : int = Query(10, ge=1, le=100),
    offset      : int = Query(0, ge=0),
    db: Session = Depends(get_db),
    current_user: TBL_AUTH_USER = Depends(require_admin),
):
    order_items = (
        db.query(TBL_ORDER_ITEM)
        .options(joinedload(TBL_ORDER_ITEM.book))
        .join(TBL_ORDER)
        .filter(func.lower(TBL_ORDER.status) == "completed")
        .all()
    )

    book_stats = {}
    for item in order_items:
        b_id = str(item.book_id) if item.book_id else "deleted"
        title = item.book.title if item.book else "Deleted Book"
        
        if b_id not in book_stats:
            book_stats[b_id] = {
                "book_title": title,
                "quantity_sold": 0,
                "total_revenue": 0.0,
                "total_profit": 0.0,
                "current_stock": item.book.stock if item.book else 0
            }
        
        sale_p = float(item.price_at_purchase or 0.0) * item.quantity
        cost_p = float(item.cost_price_at_purchase or 0.0) * item.quantity
        
        book_stats[b_id]["quantity_sold"] += item.quantity
        book_stats[b_id]["total_revenue"] += sale_p
        book_stats[b_id]["total_profit"] += (sale_p - cost_p)

    rows = list(book_stats.values())
    rows.sort(key=lambda x: x["quantity_sold"], reverse=True)

    for r in rows:
        r["total_revenue"] = round(r["total_revenue"], 2)
        r["total_profit"] = round(r["total_profit"], 2)

    return response(
        ok=True,
        status_code=status.HTTP_200_OK,
        message="Sales by book retrieved",
        data={"rows": rows}
    )